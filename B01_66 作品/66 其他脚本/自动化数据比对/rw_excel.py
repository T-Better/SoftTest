import openpyxl as opx
from config import excel_path


def ReadExcel():
    """
    aclist:ac列的list，为空时A1值为startevent1，多个值时需分别按顺序与c1组合
    allrows：每一行的所有内容，用来与数据库比对
    """
    wb1 = opx.load_workbook(excel_path)
    ws1 = wb1['Sheet1']

    # 读取AC列生成aclist
    aclist = []
    for r1 in range(2, ws1.max_row+1):
        a = ws1.cell(row=r1, column=1).value
        c = ws1.cell(row=r1, colummn=3).value
        # 如果A列为None时替换为startevent1
        if a is None:
            a = 'startevent1'
            aclist.append([a, c])
        # 如果a列内容包含多个任务（以英文逗号,分割，此时将其拆分）
        elif a.__contains__(","):
            tmpa = a.split(',')  # tmpa = ['T1','T2']
            for ai in tmpa:
                aclist.append([ai, c])
        else:
            aclist.append([a, c])

    # 按行读取所有生成的allrows
    allrows = []
    for r2 in ws1.iter_rows(min_row=2):
        allrows.append([cell.value for cell in r2])

    return aclist, allrows

