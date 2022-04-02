import openpyxl as opx
import random

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\新建100张工作表02.xlsx'
wb1 = opx.load_workbook(wp1)
# wb1.copy_worksheet(wb1['9号']) 
ws1 = wb1['9号']
# ws1.insert_rows(5,2)  # 在某个excel的第一个sheet页的第五行前面插入两行
# for i in ws1['A4':'G5']:  # 向一个区域内写入数据
#     for j in i:
#         j.value = random.randint(1,10000)

"""
获取第2列1、3、5、7行的数据
"""
# for i in range(1,8,2):
#     data1 = ws1.cell(row=i,column=2)
#     print(data1.value)

# ws1.append([i for i in range(100)])
# ws1.move_range("A1:C3",rows=2,cols=1)
# wb1.create_sheet('4月')
# ws1.move_range('A1:C3',rows=2,cols=2)
# ws1 = wb1.active
# wss1 = wb1.worksheets



wb1.save(wp1)

