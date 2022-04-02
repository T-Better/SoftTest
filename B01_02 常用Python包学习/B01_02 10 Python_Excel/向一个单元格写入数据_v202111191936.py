import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'
wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# 写入数据方法一 函数cell方法：工作表.cell(row,col,value=值)
ws1.cell(1,5,value='写入数据1')  
ws1['E2'] = '写入数据2'  # 写入数据方法二 切片赋值法：工作表[单元格]= ''
wb1.save(wp1)