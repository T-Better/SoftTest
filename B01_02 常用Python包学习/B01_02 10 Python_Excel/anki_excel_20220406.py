import openpyxl as opx

# path1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\test1644204065.xlsx'
# wb1 = opx.load_workbook(path1)
# ws1 = wb1['1号']
# 设置第一行高30，第一列宽15
# ws1.row_dimensions['A'].height = 30
# ws1.column_dimensions[1].weidth = 15
"""
# 1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
# - 按行
for j in ws1.rows(min_row=1,max_row=5,min_column='A',max_column='C'):
    for c in j:
        print(c)

# - 按列
for k in ws1.columns(min_row=1,max_row=5,min_column='A',max_column='C'):
    for c in k:
        print(c)

# - 按区域
for i in ws1['A1:C5']:
    for c in i:
        print(c)

wb1.save(path1)
"""
# 从fruit_price.xlsx中读取数据结果如下：
path2 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\fruit_price.xlsx'
wb1 = opx.load_workbook(path2)
ws1 = wb1['Sheet1']
list2 = []
for i in ws1.iter_rows(min_row=2):
    list1 = []
    for c in i:
        list1.append(c.value)
    list2.append(list1)
print(list2)


