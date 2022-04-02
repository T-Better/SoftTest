import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1["Sheet"]

# 方法一
data1 = ws1['A1':'C5']
for d in data1:
    for i in d:
        print(i.value, end=" ",)
print("/n")
# 方法二
data2 = ws1.iter_rows(min_row=1, max_row=5, min_col = 1, max_col = 3)
for d in data2:
    for i in d:
        print(i.value, end = "|")
print("/n")
# 方法三
data3 = ws1.iter_cols(min_row = 1, max_row = 5, min_col = 1, max_col = 5)
for d in data3:
    for i in d:
        print(i.value, end = "||")
print("/n")
wb1.close()