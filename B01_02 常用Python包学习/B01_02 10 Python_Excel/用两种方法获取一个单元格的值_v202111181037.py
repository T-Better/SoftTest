import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

for i in range(1,8,2):
    data1 = ws1.cell(row=i,column=2).value
    print(i,data1)

# 方法一
cell1 = ws1["A1"].value
print(cell1)

# 方法二
cell2 = wb1['Sheet']["A1"].value
print(cell2)

# 方法三
cell3 = ws1.cell(row=1,column=1).value
print(cell3)

wb1.close()
