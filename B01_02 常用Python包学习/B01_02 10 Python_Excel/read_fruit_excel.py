import openpyxl as opx


path1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\002 Docs\fruit_price.xlsx'

wb1 = opx.load_workbook(path1)
ws1 = wb1['Sheet1']

my_list = []
maxrows = ws1.max_row
for i in range(maxrows-1):  # 0-4 共5行，遍历5次
    temp_list = []  # 用来存放每行数据
    for each in ws1.iter_cols(min_row=2):  # 按列遍历，从第二行开始返回三列
        temp_list.append(each[i].value)  # 获取每一列的值第一个 第二个值放入temp_list
    my_list.append(temp_list)  # 每做成一列，加入到总列表my_list中

print(my_list)


# wb1.save(path1)





