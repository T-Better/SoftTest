import openpyxl as opx

p1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\{}'.format('test1654591815.xlsx')
wb1 = opx.load_workbook(p1)
ws1 = wb1['第1天']
ws1.merge_cells('A3:B3')
ws1.unmerge_cells('C1:D1')

wb1.save('./test20220811.xlsx')



