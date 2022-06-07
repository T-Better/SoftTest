import openpyxl as opx
"""
1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
"""
p1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\{}'.format('test1653298329.xlsx')
wb1 = opx.load_workbook(p1)
ws1 = wb1['9号']
"""
# 按行
for i in ws1.iter_rows(min_row=1,max_row=2,min_col=1,max_col=10):
    for c in i:
        print(c.value, end=' ')
print('/n')
# 按列
for j in ws1.iter_cols(min_row=1,max_row=2,min_col=1,max_col=10):
    for c in j:
        print(c.value, end='|')
print('/n')
# 按区域
for k in ws1['A1':'J3']:
    for c in k:
        print(c.value, end = "||")
"""
# 获取每一行
for i in ws1.rows:
    for c in i:
        print(c.value, end=" ")
# 每一列
for j in ws1.columns:
    for c in j:
        print(c.value, end=" | ")

wb1.close()
