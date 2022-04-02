import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# for r in ws1["A1":"C3"]:
#     for c in r:
#         c.value = "我在那"
# wb1.save(wp1)

# 向一个单元格写入数据
# ws1.cell(27,1,value = "向一个单元格写入数据")
# ws1['A28'] = '向2'
# wb1.save(wp1)

# 查询第二行第四列的位置结果，行返回2，列返回D
wcn = opx.utils.get_column_letter(2)
ncw = opx.utils.column_index_from_string('D')
print(wcn,ncw)

wb1.close()
