import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test1637828640.xlsx'

wb1 = opx.load_workbook(wp1)
ws3 = wb1['3号']
# 做批注
note1 = opx.comments.Comment('已完成','lzj')
ws3.comment = note1

# 向一个区域内写入数据
for i in ws3['A11:J20']:
    for c in i:
        c.value = "向一个区域内写入数据"

# 设置第一行高30，第一列宽15
ws3.row_dimensions[1].height = 40
ws3.column_dimensions[1].width = 15

# 删除除9号外的工作表
# wss1 = wb1.worksheets
# for i in wss1:
#     if i.title != '9号':
#         wb1.remove(i)
#     else:
#         continue

"""
对工作表中数据实现以下效果：第一行均为双实线对象，第二行及一下所有
内部边框均为单实线黑色
"""
side_double = opx.styles.Side(side='double',color = '000000')
side_thin = opx.styles.Side(side='thin',color = '000000')

border_double = opx.styles.Border(left=side_double,right=side_double,top=side_double,bottom=side_double)
border_thin = opx.styles.Border(left=side_thin,right=side_thin,top=side_thin,bottom=side_thin)

for i in ws3['A1:J10']:
    for c in i:
        c.border = border_thin

for j in ws3['A1:J1']:
    for c in j:
        c.border = border_double

# 设置第一行高30，第一列宽15
# ws3.row_dimensions[1].height = 30
# ws3.column_dimensions[1].width = 15

# 让标题水平垂直居中且自动换行,让正文自动换行
ali1 = opx.styles.Alignment(horizontal='center',vertical='center',wrap_text = true)
for k in ws3['A1:J1']:
    for c in k:
        c.alignment = ali1

ali2 = opx.styles.Alignment(horizontal='general',wrap_text = True)
for l in ws3['A2:J10']:
    for c in l:
        c.alignment = ali2

# 将模板批量复制 将Sheet1复制31份，然后删除Sheet1模板
for i in range(1,31):
    wsn = wb1.copy_worksheet(ws3)
    wsn.title = str(i) + '号'
wb1.remove(wb1['3号'])

# 标题加粗 22号 微软雅黑；正文：宋体 11号；重点：红色加粗倾斜单下划线11号字体
font_header = opx.styles.Font(name=u'微软雅黑',size=22,bold=True)
font_content = opx.styles.Font(name=u'宋体',size = 11)
font_special = opx.styles.Font(color='FF0000',bold=True,underline='single',italic=True)

for m in ws3['A1:J1']:
    for c in m:
        c.font = font_header
    
for n in ws3['A2:J10']:
    for c in n:
        c.font = font_content

for o in ws3['A3:J3']:
    for c in o:
        c.font = font_special

# 冻结test.xlsx的Sheet1的第一行
ws3.freeze_panes = 'A2'

# 在某个excel的第一个sheet页的第五列前面插入两行
ws3.insert_cols(5,2)

# 在某个excel的第一个sheet页第五列后面删除1列
ws3.delete_cols(5,1)

# 新建100张工作表 1-100号
for i in range(1,101):
    wb1.create_sheet(str(i)+'号')
    
# 移动1-3列1-3行向下移动两行，向右移动一列
ws3.move_range("A1:C3",rows=10,cols=10)

# 新建指定工作表4月
wb1.create_sheet('4月')

# 合并A3-B3单元格，取消合并C1-D1单元格
ws3.merge_cells('A3:B3')
ws3.unmerge_cells('C1:D1')

# 读取工作表S1的最大行、列、A1单元格的值（两种方法）、A1单元格的行、列位置
print(ws3.max_row)
print(ws3.max_column)

# 求第三行所有数字之和并将结果写入第三行最后一列
sum1 = "=sum(A2:J2)"
ws3['M2'].value = sum1

# 查询第二行第四列的位置结果，行返回2，列返回D
ncw = opx.utils.get_column_letter(2)

wcn = opx.utils.column_index_from_string('D')
print(ncw,wcn)

wb1.save(wp1)
