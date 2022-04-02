import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# 获取第2列1、3、5、7行的数据
for i in range(1,8,2):
    print(i, ws1.cell(row=i, column=2).value)

# 分别冻结test.xlsx的Sheet1的第一行和第一列
ws1.freeze_panes = "A2"  # 只能冻结一行或者一列，不能同时冻结

# 方法一：从sheet中直接获取单元格值
cell1 = ws1['A1'].value  
print(cell1)

# 方法二：从工作簿中直接获取单元格值
cell2 = wb1['Sheet']['A1'].value 

# 方法三：使用工作簿.cell(row=i,column=j).value方法
cell3 = ws1.cell(row=1,column=1).value 
wb1.save(wp1)

