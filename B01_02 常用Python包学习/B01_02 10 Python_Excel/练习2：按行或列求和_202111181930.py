import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']
max_rn = ws1.max_row
max_cn = ws1.max_column
for r in ws1.iter_rows(1,max_rn,1,max_cn):
    print([ce.value for ce in r])

    