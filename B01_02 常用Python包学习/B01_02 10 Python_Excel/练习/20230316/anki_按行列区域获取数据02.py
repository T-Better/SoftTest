"""
1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
"""

import openpyxl as opx


fp = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\anki_1678953594.xlsx'

wb1 = opx.load_workbook(fp)
ws1 = wb1['1号']

# 按区域
for a in ws1['A1':'C5']:
    for c in a:
        print(c.value, end='|')
    print("")
print("")

# 按行
for r in ws1.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
    for c in r:
        print(c.value,end='_')
    print("")
print("")

# 按列
for l in ws1.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    for c in l:
        print(c.value, end=' ')
    print("")
print("")

wb1.close()









