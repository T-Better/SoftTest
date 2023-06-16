# 给一下目录中的某个excel单元格A1做“已完成”批注
import openpyxl as opx

p = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
f1 = r'\anki_1678953594.xlsx'
f2 = r'\anki_设置行高列高_20230316.xlsx'

wb1 = opx.load_workbook(p + f1)
ws2 = wb1['2号']

c1 = opx.comments.Comment('已完成','chase')
ws2['A1'].comment = c1

# 合并A3-B3单元格，取消合并C1-D1单元格
ws2.merge_cells('A3:B3')

ws2.unmerge_cells('C1:D1')

wb1.save(p + f1)

