import openpyxl as opx

p = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
f1 = r'\anki_1678953594.xlsx'
f2 = r'\anki_设置行高列高_20230316.xlsx'

wb1 = opx.load_workbook(p + f1)
ws2 = wb1['2号']

# 复制指定工作表9月
wsi = wb1.copy_worksheet(ws2)
wsi.title='2yue'

# 合并A3-B3单元格，取消合并C1-D1单元格
ws2.merge_cells('A1,B3')
ws2.unmerge_cells('C1,D1')

# 求第三行所有数字之和并将结果写入第三行最后一列
ws2['J3'].value = '=sum(A3,F3)'

wb1.save(p + f1)
