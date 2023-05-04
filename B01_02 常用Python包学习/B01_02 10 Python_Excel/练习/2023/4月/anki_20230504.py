# 给一下目录中的某个excel单元格A1做“已完成”批注
import openpyxl as opx

p = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
f1 = r'\anki_1678953594.xlsx'
f2 = r'\anki_设置行高列高_20230316.xlsx'

wb1 = opx.load_workbook(p + f1)
ws2 = wb1['2号']

c1 = opx.styles.Comment('已完成','lzj')

ws2['C1'].comment = c1

# 合并A3-B3单元格，取消合并C1-D1单元格
ws2.merge_cells('C1:D1')

# 两个方法向指定一个单元格写入数据
ws2['A2'] = '1'
ws2.cell(1,5,value='写入数据1')

# 求第三行所有数字之和并将结果写入第三行最后一列
ws2['J3'] = "=SUM(A3,H3)"

# 复制指定工作表9月
wsi = wb1.copy_worksheet(wb1['1号'])
wsi.title='1号复制'

wb1.save(p + f1)









