
import openpyxl as opx

p = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
f1 = r'\anki_1678953594.xlsx'
f2 = r'\anki_设置行高列高_20230316.xlsx'

wb1 = opx.load_workbook(p + f1)
ws2 = wb1['2号']
# 按行 按列取数据
for r in ws2.iter_rows(min_row = 1):
    pass

for c in ws2.iter_cols(min_row = 1):
    pass
wb1.save(p+f1)





















