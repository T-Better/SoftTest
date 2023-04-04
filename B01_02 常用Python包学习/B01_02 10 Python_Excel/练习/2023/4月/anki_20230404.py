import openpyxl as opx

import openpyxl as opx

p = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
f1 = r'\anki_1678953594.xlsx'
f2 = r'\anki_设置行高列高_20230316.xlsx'

wb1 = opx.load_workbook(p + f1)
ws2 = wb1['2号']
# 按行
for r in ws2.iter_rows():
    pass
# 按列取数据
for l in ws2.iter_cols():
    pass

# 复制指定工作表9月
wsi = wb1.copy_worksheet(wb1['3号'])
wsi.title = ''

# 设置第一行高30，第一列宽15
ws2.row_dimensions[1].height=30
ws2.column_dimensions['A'].width = 15

# 冻结test.xlsx的Sheet1的第一行
ws2.freeze_panes = 'A2'

# 第一行 双实线 黑色
sh = opx.styles.Side(style='double',color='000000')
bh = opx.styles.Border(top=sh, bottom=sh, left=sh, right=sh)
for h in ws2['A1':'J1']:
    for c in h:
        pass  # TODO

# 标题水平垂直居中且自动换行
h = opx.styles.Alignment(horizontal = 'center',vertical='center', wrap_text=True)
for h in ws2['A1':"J1"]:
    for c in h:
        pass

wb1.save(p+f1)
