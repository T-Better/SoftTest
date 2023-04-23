import openpyxl as opx

p = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
f1 = r'\fruit_price.xlsx'

wb1 = opx.load_workbook(p + f1)
ws1 = wb1['Sheet1']

# 合并A3-B3单元格，取消合并C1-D1单元格
ws1.merge_cells('A3,B3')
ws1.unmerge_cells('C1,D1')

wb1.save(p+f1)
