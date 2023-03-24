import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'
wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

mr = ws1.max_row  # 最大行
mc = ws1.max_column  # 最大列
A1 = ws1['A1'].value  # A1单元格的值-方法一
A1_2 = ws1.cell(1,1).value  # A1单元格的值-方法二
A1_C = ws1['A1'].column  # A1的列
A1_R = ws1['A1'].row  # A1的行

# 获取C列所有数据
list1 = []
for i in ws1['C']:
    list1.append(i.value)
print(list1,end=" ")
print("/n")

# 获取第一行所有数据
list2 = []
for i in ws1[1]:
    list2.append(i.value)
print(list2, end="|")
print("/n")

# 获取所有数据
list3 = []
for i in ws1.rows:
    for j in i:
        list3.append(j.value)
print(list3,end="||")