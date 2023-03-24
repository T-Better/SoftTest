import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# 按区域
for r in ws1['A1':'C5']:
    for c in r:
        print(c.value, end=" ")
print('/n')

# 按行
for r in ws1.iter_rows(min_row=1,max_row= 5,min_col=1,max_col=3):
    for c in r:
        print(c.value, end='|')
print('/n')
# 按列
for c in ws1.iter_cols(min_row=1,max_row=5,min_col=1,max_col = 3):
    for ce in c:
        print(ce.value, end = '||')
print('/n')

wb1.close()