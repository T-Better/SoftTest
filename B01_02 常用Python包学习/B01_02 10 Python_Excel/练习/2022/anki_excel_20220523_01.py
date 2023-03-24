import openpyxl as opx
import datetime as dt
"""
# 两种方法获取工作簿test.xlsx中所有的工作表的表名
p1 = r'D:\Git_Reps\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\test.xlsx'
w1 = opx.load_workbook(p1)
wss = w1.worksheets
for i in wss:
    print(i.title)
"""
"""
# 除了9号的工作表以外都删除
p2 = r'D:\Git_Reps\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\test1653298766.xlsx'
wb2 = opx.load_workbook(p2)
wss2 = wb2.worksheets
for i in wss2:
    if i.title != '9号':
        wb2.remove(i)

wb2.save(p2)
"""

# 新建100张工作表 1-100号
p3 = r'D:\Git_Reps\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\anki_test_202205231751.xlsx'

wb3 = opx.Workbook(p3)
for i in range(100):
    wb3.create_sheet("{}号".format(i))

wb3.save(p3)









