import openpyxl as opx

wpf = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\刘子君 成方金信工作量表.xlsx'
wb1 = opx.load_workbook(wpf)
ws1 = wb1['5月']
mr = ws1.max_row
mc = ws1.max_column
A1 = ws1['A1'].value
A1_2 = ws1.cell(1,1).value
A1_C = ws1['A1'].column
A1_R = ws1['A1'].row

# 获取C列所有数据
list1 = []
for i in ws1['C']:
    list1.append(i.value)
# 获取第1行所有数据
list2 = []
for i in ws1[1]:
    list2.append(i.value)
# 获取所有数据
list3 = []
for i in ws1.rows:
    for j in i:
        list3.append(j.value)
print(list3)