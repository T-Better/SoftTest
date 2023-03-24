# anki_复习_20220207.py
import openpyxl as opx

path1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\002 Docs\test1644204065.xlsx'
wb1 = opx.load_workbook(path1)
ws1 = wb1['1号']
"""
# 冻结test.xlsx的Sheet1的第一行
# ws1.freeze_panes = 'A2'

# 合并A3-B3单元格，取消合并C1-D1单元格
# ws1.merge_cells('A3:B3')
# ws1.unmerge_cells('C1:D1')
"""

"""
对工作表中数据实现以下效果：第一行均为双实线对象，
第二行及一下所有内部边框均为单实线黑色

side_head = opx.styles.Side(style='double',color='000000')
side_content = opx.styles.Side(style='thin',color='000000')

border_head = opx.styles.Border(top='side_head',bottom='side_head',left='side_head',right='side_head')
border_content = opx.styles.Border(top='side_head',bottom='side_content',left='side_content',right='side_content')

for i in ws1['A1:J1']:
    for c in i:
        c.border = border_head

for j in ws1['A2:J10']:
    for c in j:
        c.border = border_content
"""

# 两个方法向指定一个单元格写入数据
ws1['A1'].value="1"
ws1.cell(row='1',column='A',value='book')
wb1.save(path1)
