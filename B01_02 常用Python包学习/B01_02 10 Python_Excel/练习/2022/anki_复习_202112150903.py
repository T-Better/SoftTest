import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test1637828640.xlsx'
wb1 = opx.load_workbook(wp1)
ws6 = wb1['6号']

# 将模板批量复制 将Sheet1复制31份
for i in range(1,32):
    wb1.copy_worksheet(wb1['6号'])
# 然后删除Sheet1模板
wb1.remove(wb1['6号'])

# 读取工作表S1的最大行、列
mr = ws6.max_row  # 最大行
mc = ws6.max_column # 最大列

# A1单元格的值（两种方法）
A1 = ws6['A1'].value

A_1 = ws6.cell(row=1,column="A").value

# A1单元格的行、列位置
A1_C = ws6['A1'].column  # A1的列
A1_R = ws6['A1'].row  # A1的行

wb1.save(wp1)