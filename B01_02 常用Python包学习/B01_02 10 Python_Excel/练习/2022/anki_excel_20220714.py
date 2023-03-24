"""
1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
难点：1.按行、列所用函数方法。2.按区域的参数配置方法
"""
import openpyxl as opx

path1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\test1654591815.xlsx'

wb1 = opx.load_workbook(path1)
ws1 = wb1['第1号']
"""
# 按区域
for i in ws1['A1':'C5']:
    for c in i:
        print(c.value)

# 按行
for j in ws1.iter_rows(min_row=1,max_row=5,min_col=1,max_col=3):
    for c in j:
        print(c.value)
"""

# 按列
for k in ws1.iter_cols(min_row=1,max_row=5,min_col=1,max_col=3):
    for c in k:
        print(c.value)

wb1.close()




