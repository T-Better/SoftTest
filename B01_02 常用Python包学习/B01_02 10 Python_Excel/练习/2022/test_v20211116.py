import os
import openpyxl
cwp = 'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel'  # 当前工作路径
os.chdir(cwp)
print(os.getcwd())

# 创建工作簿
wb1 = openpyxl.Workbook()  # 调用工作簿类并将其实例化
wb1.save('test_v202111161851.xlsx')  # 保存为xlsx文件

# 打开上面创建好的工作簿
wb2 = openpyxl.load_workbook('test_v202111161851.xlsx')
ws_name = wb2.sheetnames  # 获取所有工作表名称
print(ws_name)

# 选中对应的工作表
ws1 = wb2['Sheet']
print(ws1)

# 显示工作簿中所有的工作表和表名
wss = wb2.worksheets
for ws in wss:
    print(ws.title)


# 新建指定工作表 新建一个3月工作表
wb2.create_sheet('3月')


# 复制表3月成4月
cs2 = wb2.copy_worksheet(wb2['3月'])  # 这里是将要被复制的表赋值给一个对象cs2
cs2.title = '4月'


# 删除指定工作表
ws3 = wb2['3月']
wb2.remove(ws3)
wb2.save('test_v202111161851_new.xlsx')