import openpyxl as opx


path1 = ''
wb1 = opx.load_workbook(path1)
ws1 = wb1['1号']

ws1.move_range("A1:C3", rows=2, cols=1)


# 获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
# ws1["A1":"E3"]

ws1.iter_rows(min_row=1,max_row=5,min_col=1,max_col=3)

ws1.iter_cols(min_row=1,max_row=5,min_col=1,max_col=3)

wb1.save(path1)
