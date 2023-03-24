import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'
wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# 区域法
arda = ws1["A1":"E3"]  # ((第一行整体对象),(第二行整体对象))
for c in arda:  # c:一行整体 属性对象
    for j in c:  # j：一行中的一个单元格 属性对象
        print(j.value,end=" ")  # # j.value：一个单元格对应的数值
print('/n')
# 按行
for r in ws1.iter_rows(min_row=1,max_row=10,min_col=1,max_col=3):
    for c in r:
        print(c.value,end="|")
print('/n')
# 按列
for c in ws1.iter_cols(min_row=1,max_row=10,min_col=1,max_col=3):
    for e in c:
        print(e.value,end="||")

wb1.close()