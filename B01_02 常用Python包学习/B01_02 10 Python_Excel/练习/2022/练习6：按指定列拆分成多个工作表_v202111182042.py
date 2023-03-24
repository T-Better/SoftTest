import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']
wa1 = ws1.iter_rows(min_row=2)
dic1 = {}
for r in wa1:
    ed = [ce.value for ce in r] # 每行的单元格数据放到一个列表里
    if ed[1] in dic1.keys():
        dic1[ed[1]] += [ed]
    else:
        dic1[ed[1]] = [ed]
for k,v in sorted(dic1.items()):
    nws = wb1.create_sheet
    nws.append(['姓名','班级','分数'])
    for da in v:
        nws.append(da)
wb1.remove(wb1['Sheet'])
wb1.save(wp1)