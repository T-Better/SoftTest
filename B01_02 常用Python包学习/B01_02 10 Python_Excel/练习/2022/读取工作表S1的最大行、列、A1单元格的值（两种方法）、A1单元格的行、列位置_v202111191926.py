import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'
wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# 最大行
mr = ws1.max_row

# 最大列
mc = ws1.max_column

A1r = ws1['A1'].row
print(A1r)

A1c = ws1['A1'].column
print(A1c)
wb1.close()
