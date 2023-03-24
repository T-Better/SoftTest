import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test202111231100.xlsx'
wb1 = opx.load_workbook(wp1)
ws1 = wb1['2号']

# 1.让标题水平垂直居中且自动换行 2.让正文自动换行
ali1 = opx.styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
ali2 = opx.styles.Alignment(wrap_test=True)

for i in ws1['A1:J1']:
    for c in i:
        c.alignment=ali1

for j in ws1['A2:J10']:
    for c in j:
        c.alignment = ali2

# 求第三行所有数字之和并将结果写入第三行最后一列
sum3 = '=sum(A3:J3)'

# 合并A3-B3单元格，取消合并C1-D1单元格
ws1.merge_cells('A3:B3')
ws1.unmerge_cells('C1:D1')

# 冻结test.xlsx的Sheet1的第一行
ws1.freeze_panes = 'A2'

# 对工作表中数据实现以下效果：第一行均为双实线对象，第二行及一下所有内部边框均为单实线黑色
side_header = opx.styles.Side(style='double',color='000000')
side_content = opx.styles.Side(style='thin',color='000000')

border_header = opx.styles.Border(top=side_header,bottom=side_header,left=side_header,right=side_header)
border_content = opx.styles.Border(top=side_content,bottom=side_content,left=side_content,right=side_content)

for i in ws1['A1:J1']:
    for c in i:
        c.border = border_header

for j in ws1['A2:J10']:
    for c in j:
        c.border = border_content

# 冻结excel的第一行
ws1.freeze_panes = 'A2'

# 1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
  # 方法一
for i in ws1['A1:C5']:
    for c in i:
        print(c.value)

  # 方法二
for j in ws1.iter_rows(min_row=1,max_row=5,min_column=1,max_column=3):
    for c in j:
        print(c.value)

  # 方法三
for k in ws1.iter_cols(min_row=1,max_row=5,min_column=1,max_column=3):
    for c in k:
        print(c.value)

wb1.save(wp1)

