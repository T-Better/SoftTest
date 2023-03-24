import openpyxl as opx


e1 = r'\anki_1665474080.xlsx'
p1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
ep1 = p1+e1
print(ep1)
wb1 = opx.load_workbook(ep1)
ws1 = wb1['1号']

# 对工作表中数据实现以下效果：第一行均为双实线对象，第二行及一下所有内部边框均为单实线黑色
"""
line_double = opx.styles.Side(style='double',color='000000')
line_thin = opx.styles.Side(style='thin',color='000000')

border_header = opx.styles.Border(left=line_double,right=line_double,top=line_double,bottom=line_double)
border_content = opx.styles.Border(left=line_thin,right=line_thin,top=line_thin,bottom=line_thin)

for i in ws1['A2':'J10']:
    for c in i:
        c.border = border_content

for h in ws1['A1':'J1']:
    for c in h:
        c.border = border_header
"""
# 移动1-3列1-3行向下移动两行，向右移动一列
"""
ws1.move_range('A1:C3',rows=2,cols=1)
"""

# 标题加粗 22号 微软雅黑；正文：宋体 11号；重点：红色加粗倾斜单下划线11号字体
"""
font_h = opx.styles.Font(name=u'微软雅黑',bold=True,size=22)
font_c = opx.styles.Font(name=u'宋体',size=11)
font_s = opx.styles.Font(size=11,color='FF0000',bold=True,italic=True)

for r in ws1['A1:J1']:
    for c in r:
        c.font = font_h
for r in ws1['A2:J10']:
    for c in r:
        c.font = font_c
for r in ws1['A3:J3']:
    for c in r:
        c.font = font_s
"""

# 冻结excel的第一行
"""
ws1.freeze_panes='A2'
"""
# wb1.save(ep1)
"""
e2 = r'\fruit_price.xlsx'
ep2 = p1+e2
wb2 = opx.load_workbook(ep2)
ws2 = wb2['Sheet1']
res_list = []
for r in ws2.iter_rows(min_row=2):
    tmp = []
    for c in r:
        tmp.append(c.value)
    res_list.append(tmp)
print(res_list)
"""

# 在某个excel的第一个sheet页第五列后面删除1列
"""
ws1.delete_cols(5,1)
wb1.save(ep1)
"""

# 除了9月份的工作表以外都删除
"""
wss = wb1.worksheets()
for i in wss:
    if i.name != '9号':
        wb1.remove(i)
    else:
        continue
wb1.save(ep1)
"""

# 求第三行所有数字之和并将结果写入第三行最后一列
"""
list_s = []
for i in ws1['A3:J3']:
    for c in i:
        list_s.append(c.value)
print(sum(list_s))
# 法2
ws1['B6'] = '=sum(C5:O5)'
"""

# 批量修改工作表的名称 将新建100张工作表01.xlsx中第1-100天工作表名改为1-100月
"""
wss = wb1.worksheets()
for s in wss:
    s.title ="北京第" + s.title
wb1.save(ep1)
"""

# 1.让标题水平垂直居中且自动换行
# 2.让正文自动换行
"""
ali_h = opx.styles.Alignment(horizontal='center',vertical='center',wrap_text = True)
ali_c = opx.styles.Alignment(wrap_text = True)

for r in ws1['A1:J1']:
    for c in r:
        c.alignment = ali_h
for l in ws1['A2:J10']:
    for c in l:
        c.alignment = ali_c
wb1.save(ep1)
"""

# 新建指定工作表20221010
"""
wb1.create_sheet('20221010')
wb1.save(ep1)
"""

# 新建100张工作表 1-100号
for i in range(1,101):
    wb1.create_sheet(str(i)+'号')

wb1.save(ep1)



