import openpyxl as opx
import re

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test1637828640.xlsx'

wb1 = opx.load_workbook(wp1)
ws4 = wb1['4号']

ws4.move_range("A1:C3",rows=2,cols=1)

# 在某个excel的第一个sheet页第五列后面删除1列
ws4.delete_cols("C1",1)

# 设置第一行高30，第一列宽15
ws4.row_dimensions[1].height = 40
ws4.column_dimensions['A'].height = 15

# 第一行均为双实线对象，第二行及一下所有内部边框均为单实线黑色
side_double = opx.styles.Side(style='double',color='000000') # 
side_thin = opx.styles.Side(style='thin',color='000000')

border_header = opx.styles.Border(top='side_double',bottom='side_double',left='side_double',right='side_double')
border_content=opx.style.Border(top='side_thin',bottom='side_thin',left='side_thin',right='side_thin')

for i in ws4['A1:J1']:
    for c in i:
        c.border = border_header

for j in ws4['A2:J10']:
    for c in j:
        c.border = border_content
# 获取第2列1、3、5、7行的数据
for k in range(1,8,2):
    # print(ws4['k:2'].value)
    print(ws4.cell(row=i,column=2).value)

# 标题加粗 22号 微软雅黑；正文：宋体 11号；重点：红色加粗倾斜单下划线11号字体
font_header = opx.styles.Font(name=u"微软雅黑",size=22,bold=True)
for i in ws4['A1:J1']:
    for c in i:
        c.font = font_header

font_content = opx.styles.Font(name=u"宋体",size=11)
font_remark = opx.styles.Font(bold=True,color="FF0000",underline = "single",italic=True)

# 1.让标题水平垂直居中且自动换行；2.让正文自动换行
ali1 = opx.styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
for i in ws4['A1:J1']:
    for c in i:
        c.alignment = ali1

ali2 = opx.styles.Alignment(wrap_text=True)

# 批量修改工作表的名称 将新建100张工作表01.xlsx中第1-100天工作表名改为1-100月
wss1 = wb1.worksheets
for i in wss1:
    i.title=""

# 给单元格A1做“已完成”批注
note1 = opx.comments.Comment("已完成","lzj")
ws4['A1'].comment = note1

# 合并A3-B3单元格，取消合并C1-D1单元格
ws4.merge_cells('A3:B3')
ws4.unmerge_cells('C1:D1')

# 求第三行所有数字之和并将结果写入第三行最后一列
sum1 = "=SUM(A3:J3)"
ws4['G3'].value = sum1

# 获取每一行、每一列
for i in ws4.rows:
    for c in i:
        print(c.value)

for j in ws4.columns:
    for c in j:
        print(c.value)

wb1.save(wp1)
