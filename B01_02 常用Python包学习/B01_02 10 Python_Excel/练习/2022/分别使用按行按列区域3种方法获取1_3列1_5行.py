import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test1637828640.xlsx'

wb1 = opx.load_workbook(wp1)
ws3 = wb1['3号']

# 按区域
for i in ws3['A1:C5']:
    for c in i:
        print(c.value, end="|")

# for j in ws3.iter_rows(min_row=1,max_row=5,min_col=1,max_col=3):
#     for c in j:
#         print(c.value)

# for k in ws3.iter_cols(min_col=1,max_col=3,min_row=1,max_row=5):
#     for c in k:
#         print(c.value)

# 在某个excel的第一个sheet页的第五行前面插入两行
# ws3.insert_rows(5,2)

# 设置第一行高30，第一列宽15
# ws3.row_dimensions[1].height = 30
# ws3.column_dimensions["A"].width = 15

# 让标题水平垂直居中且自动换行,让正文自动换行
# ali_header = opx.styles.Alignment(horizontal="center",vertical="center",wrap_text=True)
# ali_content = opx.styles.Alignment(horizontal="center",wrap_text=True)

# for i in ws3["A1:J1"]:
#     for c in i:
#         c.alignment = ali_header

# for j in ws3['A2:J10']:
#     for c in j:
#         c.alignment = ali_content

# 给单元格A1做“已完成”批注
note1 = opx.comments.Comment('已完成','lzj')
ws3['A2'].comment = note1

wb1.save(wp1)