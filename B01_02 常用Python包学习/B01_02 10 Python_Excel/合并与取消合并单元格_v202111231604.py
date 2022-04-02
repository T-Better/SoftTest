import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test202111231100.xlsx'
wb1 = opx.load_workbook(wp1)
ws1 = wb1['1月']

ws1.merge_cells('A3:B3')  # 合并A3-B3单元格
ws1.unmerge_cells('C1:D1')  # 取消合并C1-D1单元格

wb1.save(wp1)