import openpyxl as opx


wdf1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\新建100张工作表 1-100号1118.xlsx'
wb1 = opx.Workbook(wdf1)  # 创建工作簿
# wb1.save(wdf1) 有这个没这个区别就是有的话会默认自动创建一个Sheet1，没有不会创建

# 加载工作簿
# wb1 = opx.load_workbook(wdf1)
# 创建100张工作表
for i in range(1,101):
    ws1 = wb1.create_sheet(str(i) + "号")
# 查看结果
wss1 = wb1.sheetnames
print(wss1)

# 完事记得保存
wb1.save(wdf1)
