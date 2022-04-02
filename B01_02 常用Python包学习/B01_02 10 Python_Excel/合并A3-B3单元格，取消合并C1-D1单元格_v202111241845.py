import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test1637753231.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['1号']

ws1.merge_cells('A1:B3')
ws1.unmerge_cells('C1:D1')

wb1.save(wp1)