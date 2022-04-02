import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test1640048977.xlsx'
wb1 = opx.load_workbook(wp1)
ws2 = wb1['2号']

# 在某个excel的第一个sheet页第五列后面删除1列
ws2.delete_cols(5,1)

# 两个方法向指定一个单元格写入数据
ws2['A1'].value = 'zhangsan'
ws2.cell(row=1,column='A',value='lisi')

# 查询第二行第四列的位置结果，行返回2，列返回D
cd = opx.utils.get_column_letter(2)


