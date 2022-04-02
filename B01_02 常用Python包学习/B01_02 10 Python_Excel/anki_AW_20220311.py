import openpyxl as opx

path1 = r'D:\BaiduNetdiskWorkspace\Super_coder\Soft_test\C66_03 自动化办公\C06_03 02 Excel_Python\002 Docs\fruit_price.xlsx'

# wb1 = opx.Workbook(path1)
# for i in range(1,101):
#     wb1.create_sheet("{}号".format(i))
#     wb1.save(path1)

wb1 = opx.load_workbook(path1)
ws1 = wb1['Sheet1']

ws1.freeze_panes = 'B1'

list1 = []
for row in ws1.iter_rows(min_row=2):
    list2 = []
    for c in row:
        list2.append(c.value)
    list1.append(list2)

print(list1)