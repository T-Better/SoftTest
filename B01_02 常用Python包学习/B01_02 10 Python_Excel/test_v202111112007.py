import os,openpyxl
from openpyxl import Workbook

os.chdir('D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel')

# 1.创建工作簿
# wb = Workbook()
# ws = wb.active


# ws1 = wb.create_sheet()  # 默认插在工作簿末尾
# ws1.title = '李明远'
# ws.sheet_properties.tabColor = '1072BA'
# wb.save('D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel')
# wb.save('roommate.xlsx')#
wb = openpyxl.load_workbook('roommate.xlsx')
ws = wb['李明远']
print(ws)