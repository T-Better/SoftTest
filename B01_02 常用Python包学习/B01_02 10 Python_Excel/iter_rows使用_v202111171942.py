import openpyxl as opx

wpf = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\刘子君 成方金信工作量表.xlsx'
wb1 = opx.load_workbook(wpf)
ws1 = wb1['5月']
# iter_cols按列
for r in ws1.iter_rows(min_row=1,max_row=10,min_col=1,max_col=3):
    for c in r:
        print(c.value)

# iter_cols按行 1-10行 1-3列
for c in ws1.iter_cols(min_row=1,max_row=10,min_col=1,max_col=3):
    for e in c:
        print(e.value)