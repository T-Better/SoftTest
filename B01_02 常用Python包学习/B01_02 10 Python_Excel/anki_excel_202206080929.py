import openpyxl as opx
import os


p1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs\{}'.format('test1654591815.xlsx')

wb1 = opx.load_workbook(p1)
ws1 = wb1[0]

# *读取工作表S1的最大行、列
mr = ws1.max_row
mc = ws1.max_column

# A1单元格的值（两种方法）
c1 = ws1['A1'].value
print(c1)

cc1 = ws1.cell(row=1,column='A')
print(cc1.value)

# A1单元格的行、列位置
A1 = ws1['A1'].column
AA1 = ws1['A1'].row
print(A1)
print(AA1)

# 复制指定工作表9月
cp1 = wb1.copy_worksheet(wb1['9月'])
cp1.title = ''

# path路径下有一个"孙兴华.xls"文件，用两种方法找到他并输出他的名字、所在路径和是否是目录
for t in os.scandir('path'):
    if t.name.startwith('孙兴华'):
        print(t.name,t.path,t.is_dir())

wb1.close()
