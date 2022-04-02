import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
ws1 = wb1['Sheet']

# ws1.append(['20','21','11','22','10','05'])

# ws1.cell(1,5,value='写入数据2')
# ws1['E2'] = '写入数据3'

# for r in ws1['A22':'C24']:
#     for c in r:
#         c.value = 220

# wb1.save(wp1)
# 查询第二行第四列的位置结果，行返回2，列返回D
# r = opx.utils.get_column_letter(2)
# print(r)

# wcn = opx.utils.column_index_from_string('D')
# print(wcn)

# ws1.cell(25,1,value='张三')
# ws1['A25'] = '数据'

# ws1.append(["11","22","33","44"])

# max_r = ws1.max_row
# print(max_r)
# wb1.save(wp1)

# max_c = ws1.max_column
# print(max_c)

wcn = opx.utils.get_column_letter(2)
print(wcn)

ncw = opx.utils.column_index_from_string('D')
print(wcn)

wb1.close()
