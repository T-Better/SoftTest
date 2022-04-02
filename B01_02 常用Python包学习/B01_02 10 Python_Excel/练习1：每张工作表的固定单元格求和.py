import openpyxl as opx

wp1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\a\test.xlsx'

wb1 = opx.load_workbook(wp1)
list1 = []
for ws in wb1.worksheets:
    if type(ws['D6'].value) == int:  # 如果D6的值是整数则放到列表中进行计算
        list1.append(int(ws['D6'].value))
    else:
        continue  # 否则下一个
print(sum(list1))

