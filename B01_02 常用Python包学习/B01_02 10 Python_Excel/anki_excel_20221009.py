import openpyxl as opx

p1 = r'D:\Git_Reps\SoftTest\Soft_test\SoftTest\B01_02 常用Python包学习\B01_02 10 Python_Excel\002 Docs'
e1 = r'\anki_1665318831.xlsx'
ep1 = p1+e1
print(ep1)
wb1 = opx.load_workbook(ep1)
"""
# 在某个excel的第一个sheet页的第五列前面插入两列
ws1 = wb1['1号']
ws1.insert_cols(5,2)
"""

"""
# 1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
# 难点：1.按行、列所用函数方法。2.按区域的参数配置方法
ws1 = wb1['1号']

# 一、按区域
# for i in ws1['A1':'C5']:
#     for c in i:
#         print(c.value)

# 二、按行
# for row in ws1.iter_rows(min_row=1,min_col=1,max_row=5,max_col=3):
#     for c in row:
#         print(c.value)

# 三、按列
# for col in ws1.iter_cols(min_row=1,min_col=1,max_row=5,max_col=3):
#     for c in col:
#         print(c.value)
"""

# 复制指定工作表9月
# wsi = wb1.copy_worksheet(wb1['1号'])
# wsi.title = '新1号'







wb1.save(ep1)





















