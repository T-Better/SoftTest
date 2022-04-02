# anki_自动化办公_20220208.py
import openpyxl as opx

path1 = r'D:\BaiduNetdiskWorkspace\Super_coder\O03_Office_Auto_My_code\Excel_001_读写excel\002 Docs\test1644204065.xlsx'
wb1 = opx.load_workbook(path1)
ws2 = wb1['2号']
"""
1.获取1-3列，1-5行数据，分别使用按行、按列、区域3种方法法
"""
# 方法一：
for i in ws2['A1:C5']:
    for c in i:
        print(c.value,end="|")
print("----------------------------------")
# 方法二：
for j in ws2.iter_rows(min_row=1,max_row=5,min_col=1,max_col=3):
    for c in j:
        print(c.value,end=" ")
print("------------------------------------")
# 方法三
for k in ws2.iter_cols(min_row=1,max_row=5,min_col=1,max_col=3):
    for c in k:
        print(c.value,end="||")
print("--------------------------------------")
wb1.save(path1)
