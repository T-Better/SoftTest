from openpyxl import load_workbook
from app import BASE_DIR
import os
"""封装读取数据文件的方法"""

class ExcelUtil(object):
    # 初始化，增加默认属性excel_path和sheet_name（默认值均为None）
    # 注意，代码中都不使用常量路径，而都使用变量路径
    def __init__(self,excel_path=None, excel_name = None, sheet_name=None):
        """获取excel工作表"""
        # TODO 9行 防止换个文件夹就跑不起来;excel_path应该带文件名，而且路径故意不传，只传excel_name 9行
        if excel_path == None:
            self.excel_path = BASE_DIR + excel_name
        else:
            self.excel_path = excel_path

        if sheet_name == None:
            self.sheet_name = sheet_name  # TODO 不会了，怎么取sheet名字？调用时候传进来的 sheet_name
        else:
            self.sheet_name = sheet_name

        # 打开工作表 2行
        self.wb1 = load_workbook(self.excel_path)  # TODO excel_path没有excel名字，图中方法加进来感觉有点不灵活:写到excel_path中
        self.ws1 = self.wb1[self.sheet_name]

    def get_data(self):
        """
        获取文件数据
        每一行数据一个list case，所有的数据一个大list case_all:[['用例编号', '标题', '邮箱地址', '密码', '预期结果定位', '预期结果'], ['login-001', '登录成功', 'admin@tynam.com', 'tynam123', 'login success', '登录成功！'], ...
        注意要判断是否没内容，没内容的话就打印 总行数小于1，没有数据  13行
        """
        row_nums = self.ws1.max_row
        col_nums = self.ws1.max_column
        if row_nums <= 1:
            print("打印总行数小于1，没有数据")
        else:
            case_all = []
            for row in self.ws1.rows:
                case = []
                for i in range(col_nums):  # for c in row方法也行
                    case.append(row[i].value)
                case_all.append(case)
            print(case_all)  # TODO 测试代码
            return case_all

if __name__ == "__main__":
    # 调用并打印出数据 3行 # TODO 下面的都没写出来
    sheetname="LoginTest02"
    excelname=r"\data\casedata2.xlsx"
    file = ExcelUtil(sheet_name=sheetname, excel_name = excelname)
    print(file.get_data())






